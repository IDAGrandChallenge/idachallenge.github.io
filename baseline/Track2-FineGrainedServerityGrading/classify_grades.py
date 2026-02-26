import os
import re
import numpy as np
import pandas as pd
import cv2
from scipy.spatial import distance_matrix
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import json 

# ===================== 全局核心配置（用户仅需修改此处）=====================

################################### 第一步配置：图片文件夹 + YOLOv5分割标签文件夹 ###################################
IMAGE_FOLDER = '/home/data/result/predict/'      # 待分级的缺陷图+标签(yolo分割的标签)
TXT_FOLDER = '/home/data/result/predict/labels/'
SAVE_FOLDER = '/home/data/result/predict_grading/' # 保存路径-保存的内容为每张缺陷图对应的注入面积、对比度等信息，用于第二步统计成一个excel的直观表格
BG_EXPAND_RATIO = 0.1  # 背景区域相对缺陷外接矩形的扩展比例

################################### 第二步配置：解析txt生成原始Excel（路径保存在YOLO_TXT_FOLDER，无需修改）###################################
# TODO 不用
ORIGINAL_EXCEL_NAME = "缺陷指标汇总表（按缺陷分行）.xlsx"  # 接第一步得到详细的缺陷统计标

################################### 第三步配置：生成分级Excel###################################
GRADED_EXCEL_NAME = "缺陷指标汇总表（带分级）.xlsx"                           #
FACTORS_CONFIG = [
    {
        'column': '损伤区面积（像素²）',
        'scores': [3, 5, 7, 9],
        'grades': ['Acceptable', 'Marginal NG', ', NG,', 'Gross NG']
    },
    {
        'column': '外接矩形长度（像素）',
        'scores': [2.5, 4.5, 6.5, 8.5],
        'grades': ['Acceptable', 'Marginal NG', ', NG,', 'Gross NG']
    },
    {
        'column': '外接矩形高度（像素）',
        'scores': [3, 5, 7, 9],
        'grades': ['Acceptable', 'Marginal NG', ', NG,', 'Gross NG']
    },
    {
        'column': '背景平均灰度值',
        'scores': [2, 4, 6, 8],
        'grades': ['Acceptable', 'Marginal NG', ', NG,', 'Gross NG']
    }
]
WEIGHTS = [0.15, 0.25, 0.25, 0.35]  # 对应上述三个因素的权重

################################### 第三步配置：生成分级Excel###################################

# ===================== 代码0：图片+YOLO标签 → defect_result.txt =====================
class YOLODefectAnalyzer:
    def __init__(self, img_folder, txt_folder, save_folder, bg_expand_ratio=0.1):
        """
        初始化YOLOv5分割标签分析器（含灰度特征计算）
        :param img_folder: 图片文件夹路径（r''格式）
        :param txt_folder: YOLOv5分割标签文件夹路径（r''格式）,默认分割标签和图片保存在同一路径
        :param save_folder: 生成结果保存路径
        :param bg_expand_ratio: 背景区域相对缺陷外接矩形的扩展比例（默认0.1）
        """
        self.img_folder = img_folder
        self.txt_folder = txt_folder
        self.save_folder = save_folder
        self.bg_expand_ratio = bg_expand_ratio  # 背景区域扩展比例
        self.image_extensions = ('.jpg', '.jpeg', '.png', '.bmp', '.tiff')
        self._validate_folders()

    def _validate_folders(self):
        if not os.path.isdir(self.img_folder):
            raise ValueError(f"图片文件夹不存在：{self.img_folder}")
        if not os.path.isdir(self.txt_folder):
            raise ValueError(f"分割标签文件夹不存在：{self.txt_folder}")

    def _get_image_data(self, img_path):
        """获取图片的灰度图、宽高"""
        try:
            with Image.open(img_path) as img:
                gray_img = img.convert('L')  # 转为灰度图
                gray_array = np.array(gray_img, dtype=np.uint8)
                width, height = img.size
                return gray_array, width, height
        except Exception as e:
            print(f"【代码0】警告：无法读取图片 {img_path}，错误：{str(e)}")
            return None, None, None

    def _parse_yolo_txt(self, txt_path, img_width, img_height):
        """解析YOLOv5格式分割标签，返回像素坐标点集"""
        defects = []
        if not os.path.exists(txt_path):
            return defects

        with open(txt_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                parts = line.split()
                if len(parts) < 5:
                    print(f"【代码0】警告：{txt_path} 标签格式错误：{line}")
                    continue

                cls = int(parts[0])
                norm_coords = list(map(float, parts[1:]))
                if len(norm_coords) % 2 != 0:
                    print(f"【代码0】警告：{txt_path} 坐标数量异常：{line}")
                    continue

                # 归一化坐标转像素坐标
                points = []
                for i in range(0, len(norm_coords), 2):
                    x_pixel = round(norm_coords[i] * img_width)
                    y_pixel = round(norm_coords[i + 1] * img_height)
                    points.append((x_pixel, y_pixel))

                if len(points) >= 2:
                    defects.append({'class': cls, 'points': np.array(points, dtype=np.int32)})
        return defects

    def _generate_defect_mask(self, points, img_shape):
        """根据点集生成缺陷区域掩码（1为缺陷，0为背景）"""
        mask = np.zeros(img_shape, dtype=np.uint8)
        # 填充多边形区域为缺陷
        cv2.fillPoly(mask, [points], 1)
        return mask

    def _calculate_defect_gray(self, gray_array, mask):
        """计算缺陷区域的平均灰度值"""
        defect_pixels = gray_array[mask == 1]  # 提取缺陷区域像素
        if len(defect_pixels) == 0:
            return None
        return round(np.mean(defect_pixels), 2)

    def _calculate_background_gray(self, gray_array, mask, points):
        """计算缺陷周边背景区域的平均灰度值"""
        # 获取缺陷外接矩形
        x_min, y_min = np.min(points, axis=0)
        x_max, y_max = np.max(points, axis=0)
        w, h = x_max - x_min, y_max - y_min

        # 扩展矩形作为背景区域（避免包含缺陷）
        expand_w = int(w * self.bg_expand_ratio)
        expand_h = int(h * self.bg_expand_ratio)
        bg_x_min = max(0, x_min - expand_w)
        bg_y_min = max(0, y_min - expand_h)
        bg_x_max = min(gray_array.shape[1] - 1, x_max + expand_w)
        bg_y_max = min(gray_array.shape[0] - 1, y_max + expand_h)

        # 生成背景区域掩码（排除缺陷区域）
        bg_mask = np.zeros(gray_array.shape, dtype=np.uint8)
        bg_mask[bg_y_min:bg_y_max + 1, bg_x_min:bg_x_max + 1] = 1  # 背景候选区
        bg_mask[mask == 1] = 0  # 剔除缺陷区域

        background_pixels = gray_array[bg_mask == 1]
        if len(background_pixels) < 10:  # 确保背景像素数量足够
            return None
        return round(np.mean(background_pixels), 2)

    def _calculate_area(self, points):
        if len(points) < 3:
            return 0.0
        if not np.array_equal(points[0], points[-1]):
            points = np.vstack([points, points[0]])
        x, y = points[:, 0], points[:, 1]
        area = 0.5 * np.abs(np.sum(x * np.roll(y, -1) - np.roll(x, -1) * y))
        return round(area, 4)

    def _calculate_max_linear_length(self, points):
        if len(points) < 2:
            return 0.0
        dist_matrix = distance_matrix(points, points)
        return round(np.max(dist_matrix), 4)

    def _calculate_min_rect(self, points):
        x_min, y_min = np.min(points, axis=0)
        x_max, y_max = np.max(points, axis=0)
        w, h = x_max - x_min, y_max - y_min
        return round(w, 2), round(h, 2)

    def process_all(self):
        """批量处理所有图片和标签，生成缺陷报告txt"""
        img_count = 0
        for img_name in os.listdir(self.img_folder):
            if img_name.lower().endswith(self.image_extensions):
                img_count += 1
                img_basename = os.path.splitext(img_name)[0]
                img_path = os.path.join(self.img_folder, img_name)
                txt_path = os.path.join(self.txt_folder, f"{img_basename}.txt")
                print(f"\n【代码0】处理文件：{img_name}")

                # 获取灰度图和尺寸
                gray_array, img_width, img_height = self._get_image_data(img_path)
                if gray_array is None:
                    continue

                # 解析缺陷标签
                defects = self._parse_yolo_txt(txt_path, img_width, img_height)
                if not defects:
                    print(f"【代码0】  未检测到有效缺陷标签：{txt_path}")
                    continue

                # 计算并保存结果
                result_path = os.path.join(self.save_folder, f"{img_basename}_defect_result.txt")
                with open(result_path, 'w', encoding='utf-8') as f:
                    f.write(f"图片：{img_name}（尺寸：{img_width}x{img_height}像素）\n")
                    f.write("=" * 70 + "\n")

                    for i, defect in enumerate(defects, 1):
                        cls = defect['class']
                        points = defect['points']
                        img_shape = gray_array.shape[:2]  # (高, 宽)

                        # 生成缺陷掩码
                        mask = self._generate_defect_mask(points, img_shape)

                        # 计算灰度特征
                        defect_gray = self._calculate_defect_gray(gray_array, mask)
                        bg_gray = self._calculate_background_gray(gray_array, mask, points)
                        gray_diff = round(defect_gray - bg_gray, 2) if (defect_gray and bg_gray) else None
                        rect_w, rect_h = self._calculate_min_rect(points)

                        # 计算几何特征
                        area = self._calculate_area(points)
                        max_len = self._calculate_max_linear_length(points)

                        # 写入结果
                        f.write(f"缺陷 {i}（类别 {cls}）：\n")
                        f.write(f"  点集数量：{len(points)} 个\n")
                        f.write(f"  损伤区面积：{area} 像素²\n")
                        f.write(f"  外接矩形长度：{rect_w} 像素\n")
                        f.write(f"  外接矩形高度：{rect_h} 像素\n")
                        f.write(f"  最长线性长度（最远点距离）：{max_len} 像素\n")
                        f.write(f"  缺陷平均灰度值：{defect_gray if defect_gray else '无法计算'}\n")
                        f.write(f"  背景平均灰度值：{bg_gray if bg_gray else '无法计算'}\n")
                        f.write(f"  缺陷-背景灰度差值：{gray_diff if gray_diff else '无法计算'}\n")
                        f.write("-" * 70 + "\n")

                print(f"【代码0】  计算完成，结果保存至：{result_path}")

        if img_count == 0:
            raise ValueError(f"【代码0】图片文件夹中未找到有效图片：{self.img_folder}")
        print(f"\n【代码0】所有文件处理完毕！共处理 {img_count} 张图片")

# ===================== 代码1：缺陷txt解析 → 原始Excel =====================
class DefectResultExcelGenerator:
    def __init__(self, result_folder):
        self.result_folder = result_folder
        self.data = []
        self.headers = [
            "图片名称",
            "缺陷序号",
            "缺陷类别",
            "损伤区面积（像素²）",
            "外接矩形长度（像素）",
            "外接矩形高度（像素）",
            "最长线性长度（像素）",
           # "最长线性长度对应点",
            "缺陷平均灰度值",
            "背景平均灰度值",
            "缺陷-背景灰度差值"
        ]
        self._validate_folder()
        self._parse_all_results()

    def _validate_folder(self):
        if not os.path.isdir(self.result_folder):
            raise ValueError(f"文件夹不存在：{self.result_folder}（请使用r''格式输入路径）")

    def _parse_single_file(self, file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()

            # 提取图片名称
            img_name_match = re.search(r'图片：(.+?)\（尺寸', content)
            if img_name_match:
                img_name = img_name_match.group(1).strip()
            else:
                img_name = os.path.basename(file_path).replace('_defect_result.txt', '').strip()

            # 提取所有缺陷块
            defect_blocks = re.findall(r'缺陷 (\d+)\（类别 (\d+)\）：(.*?)(?=缺陷 \d+|=|$)', content, re.DOTALL)

            for block in defect_blocks:
                defect_idx, defect_cls, block_content = block
                block_content = block_content.strip()

                # 提取各项指标
                area_match = re.search(r'损伤区面积：([\d.]+|无法计算)', block_content)
                area = area_match.group(1) if area_match else "N/A"

                len_match = re.search(r'外接矩形长度：([\d.]+|无法计算)', block_content)
                length = len_match.group(1) if len_match else "N/A"

                height_match = re.search(r'外接矩形高度：([\d.]+|无法计算)', block_content)
                height = height_match.group(1) if height_match else "N/A"

                max_len_match = re.search(r'最长线性长度（最远点距离）：([\d.]+|无法计算)', block_content)
                max_len = max_len_match.group(1) if max_len_match else "N/A"

                # points_match = re.search(r'最长线性长度对应点：(.+?)\n', block_content)
                # points = points_match.group(1).strip() if points_match else "N/A"

                defect_gray_match = re.search(r'缺陷平均灰度值：([\d.]+|无法计算)', block_content)
                defect_gray = defect_gray_match.group(1) if defect_gray_match else "N/A"

                bg_gray_match = re.search(r'背景平均灰度值：([\d.]+|无法计算)', block_content)
                bg_gray = bg_gray_match.group(1) if bg_gray_match else "N/A"

                diff_match = re.search(r'缺陷-背景灰度差值：([\d.-]+|无法计算)', block_content)
                gray_diff = diff_match.group(1) if diff_match else "N/A"

                # 追加单个缺陷数据
                self.data.append([
                    img_name,
                    defect_idx,
                    defect_cls,
                    area,
                    length,
                    height,
                    max_len,
                    # points,
                    defect_gray,
                    bg_gray,
                    gray_diff
                ])

        except Exception as e:
            print(f"【代码1】解析文件 {os.path.basename(file_path)} 失败：{str(e)}")

    def _parse_all_results(self):
        result_files = [
            f for f in os.listdir(self.result_folder)
            if f.endswith('_defect_result.txt') and os.path.isfile(os.path.join(self.result_folder, f))
        ]

        if not result_files:
            raise ValueError(f"【代码1】未找到defect_result.txt文件：{self.result_folder}")

        for file in result_files:
            self._parse_single_file(os.path.join(self.result_folder, file))

        print(f"【代码1】解析完成，共 {len(self.data)} 个缺陷数据")

    def generate_excel(self, output_filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "defects grading"

        # 表头样式
        header_font = Font(bold=True)
        align_center = Alignment(horizontal="center", vertical="center")

        # 写入表头
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = align_center

        # 写入数据
        for row, data_row in enumerate(self.data, 2):
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                if col in [1, 8]:
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="right")

        # 调整列宽
        col_widths = [20, 8, 8, 15, 15, 15, 15, 30, 15, 15, 18]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # 保存Excel
        output_path = os.path.join(self.result_folder, output_filename)
        wb.save(output_path)
        print(f"【代码1】原始Excel已保存至：{output_path}")
        return output_path

# ===================== 代码2：原始Excel → 带分级Excel =====================
def grade_value(value, bins, scores):
    try:
        if str(value) in ["无法计算", "N/A"]:
            return "N/A"
        for i in range(len(bins) - 1):
            if bins[i] <= float(value) < bins[i + 1]:
                return scores[i]
    except Exception as e:
        print(f"代码错误{e}")
    return scores[-1]

def get_grade_from_score(score, scores):
    grade_map = {
        scores[0]:'Acceptable',
        scores[1]: 'Marginal NG',
        scores[2]: 'NG',
        scores[3]: 'Gross NG'
    }
    return grade_map.get(score, '未知')

def calculate_weighted_score_and_grade(row, factor_columns, weights):
    weighted_score = 0.0
    for idx, col in enumerate(factor_columns):
        score_col = col + '得分'
        score_val = row[score_col]
        # 如果有非数字／非合法得分，直接返回 N/A
        if not isinstance(score_val, (int, float)):
            return pd.Series(["N/A", "N/A"], index=['加权总分', '最终等级'])
        weighted_score += row[score_col] * weights[idx]

    if weighted_score <= 3:
        final_grade = 'Acceptable'
    elif 3 < weighted_score <= 4.5:
        final_grade = 'Marginal NG'
    elif 4.5 < weighted_score <= 7.5:
        final_grade = 'NG'
    elif weighted_score > 7.5:
        final_grade = 'Gross NG'

    return pd.Series([weighted_score, final_grade], index=['加权总分', '最终等级'])

def process_excel_to_graded(input_excel, output_excel, factors_config, weights):
    # 读取Excel
    excel_file = pd.ExcelFile(input_excel, engine='openpyxl')
    sheet_names = excel_file.sheet_names

    # 提取因素列名
    factor_columns = [config['column'] for config in factors_config]

    # 写入带分级的Excel
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        for sheet_name in sheet_names:
            if sheet_name not in excel_file.sheet_names:
                print(f"【代码2】警告：输入文件中不存在子表: {sheet_name}，跳过")
                continue

            df = excel_file.parse(sheet_name)
            if df.empty or sheet_name not in excel_file.sheet_names:
                print(f"【代码2】跳过空表: {sheet_name}")
                continue

            print(f"【代码2】处理子表: {sheet_name}，共{len(df)}条数据")

            # 处理每个因素
            for config in factors_config:
                factor = config['column']
                scores = config['scores']

                # 过滤有效数值（排除无法计算、N/A）
                df_valid_factor = df[df[factor].apply(lambda x: str(x).replace('.', '').isdigit())]
                if len(df_valid_factor) == 0:
                    print(f"【代码2】警告：子表{sheet_name}的{factor}无有效数值，跳过该因素分级")
                    df[factor + '得分'] = "N/A"
                    df[factor + '等级'] = "N/A"
                    continue

                sorted_values = sorted(df_valid_factor[factor].astype(float))
                n = len(sorted_values)

                # 计算分箱
                if n >= 4:
                    bins = [
                        sorted_values[0],
                        sorted_values[n // 4],
                        sorted_values[2 * n // 4],
                        sorted_values[3 * n // 4],
                        sorted_values[-1]
                    ]
                else:
                    bins = [sorted_values[0], sorted_values[-1], sorted_values[-1], sorted_values[-1]]
                    print(f"【代码2】警告: 子表{sheet_name}的{factor}数据量不足3条，使用降级处理")

                # 计算得分和等级
                score_column = factor + '得分'
                df[score_column] = df[factor].apply(
                    lambda x: grade_value(float(x) if str(x).replace('.', '').isdigit() else x, bins, scores)
                    if str(x) not in ["无法计算", "N/A"] else "N/A"
                )

                grade_column = factor + '等级'
                df[grade_column] = df[score_column].apply(
                    lambda s: get_grade_from_score(s, scores) if s in scores else "N/A"
                )

            # 计算加权总分和最终等级
            df_valid_total = df[df[factor_columns[0] + '得分'].apply(lambda x: x in factors_config[0]['scores'])]
            if not df_valid_total.empty:
                df[['加权总分', '最终等级']] = df.apply(
                    lambda row: calculate_weighted_score_and_grade(row, factor_columns, weights),
                    axis=1
                )
            else:
                df['加权总分'] = "N/A"
                df['最终等级'] = "N/A"

            # 保存子表

            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"【代码2】子表{sheet_name}处理完成")

    print(f"【代码2】带分级Excel已保存至：{output_excel}")
    return output_excel

# ===================== 代码3：带分级Excel → 等级标注json =====================
def write_defect_severity_to_json(graded_excel, txt_dir, output_root):
    # 验证Excel
    if not os.path.exists(graded_excel):
        print(f"【代码3】❌ 读取Excel文件失败：文件不存在 → {graded_excel}")
        return

    # 读取Excel
    try:
        excel_file = pd.ExcelFile(graded_excel)
        sheet_names = excel_file.sheet_names
        print(f"【代码3】✅ 成功读取Excel文件，包含子表：{sheet_names}")
    except Exception as e:
        print(f"【代码3】❌ 读取Excel文件失败：{str(e)}")
        return

    # 遍历子表
    for sheet in sheet_names:
        folder_path = os.path.join(output_root, sheet)
        os.makedirs(folder_path, exist_ok=True)
        print(f"\n【代码3】===== 处理子表：{sheet}（输出文件夹：{folder_path}） =====")

        # 读取子表数据
        try:
            df = pd.read_excel(graded_excel, sheet_name=sheet, keep_default_na=False)
            df_valid = df.dropna(subset=["图片名称", "最终等级"])
            print(f"【代码3】子表{sheet}原始数据{len(df)}行，有效缺陷记录{len(df_valid)}行")
        except Exception as e:
            print(f"【代码3】❌ 读取子表{sheet}失败：{str(e)}")
            continue

        # 检查必要列
        required_cols = ["图片名称", "最终等级"]
        if not all(col in df.columns for col in required_cols):
            print(f"【代码3】❌ 子表{sheet}缺少必要列：{required_cols}，跳过该子表")
            continue

        # 遍历有效记录
        all_data = {}
        for idx, row in df_valid.iterrows():
            excel_row_num = idx + 2
            img_name = str(row["图片名称"]).strip()
            index = int(row["缺陷序号"])
            severity = str(row["最终等级"]).strip().replace("　", " ")

            # 等级转换
            # if raw_severity in ["serious", "moderate", "slight"]:
            #     severity = raw_severity
            # elif raw_severity in SEVERITY_MAP:
            #     severity = SEVERITY_MAP[raw_severity]
            #     print(f"【代码3】【转换】Excel行{excel_row_num}：{raw_severity} → {severity}")
            # else:
            #     cleaned_sev = raw_severity.replace(" ", "")
            #     if cleaned_sev in SEVERITY_MAP:
            #         severity = SEVERITY_MAP[cleaned_sev]
            #         print(f"【代码3】【转换】Excel行{excel_row_num}：{raw_severity} → {cleaned_sev} → {severity}")
            #     else:
            #         print(f"【代码3】【跳过】Excel行{excel_row_num}：等级{raw_severity}不合法")
            #         continue

            # 写入json
            txt_filename = os.path.splitext(img_name)[0] + ".txt"
            txt_old_path = os.path.join(txt_dir, txt_filename)
            if not os.path.exists(txt_old_path):
                print(f"not found txt path {txt_old_path}")
                continue
            try:
                data = {}
                with open(txt_old_path, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                
                line = lines[index-1].strip().split(" ")
                data = {
                    "class": line[0],
                    "points": ', '.join(line[1:]),
                    "severity":severity
                }

                if txt_filename.split(".")[0] not in list(all_data.keys()):
                    all_data[os.path.splitext(img_name)[0]] = [data]
                else:
                    all_data[os.path.splitext(img_name)[0]].append(data)
            
                # with open(txt_save_path, "a", encoding="utf-8") as f:
                #     f.write(lines[index-1])
                #     f.write(f"severity: {severity}\n")
                    
                defect_list = list(df_valid["图片名称"])[:idx+1]
                defect_idx = defect_list.count(img_name)
            except Exception as e:
                print(f"【代码3】【失败】Excel行{excel_row_num} → 错误：{str(e)}")

        for txt_filename, data in all_data.items():
            txt_save_path = os.path.join(folder_path, txt_filename+".json")
            with open(txt_save_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"【代码3】【成功】Excel行{excel_row_num} → {txt_save_path} | 图片{img_name}的第{defect_idx}个缺陷 | 等级：{severity}")


        
    print("\n【代码3】" + "="*60)
    print("【代码3】所有子表处理完成！✅")

# ===================== 端到端流程执行 =====================
if __name__ == "__main__":
    try:
        # 步骤1：执行代码0，生成defect_result.txt
        print("="*80)
        print("开始执行步骤1：图片+YOLO标签 → 生成缺陷报告txt")
        print("="*80)
        yolo_analyzer = YOLODefectAnalyzer(IMAGE_FOLDER, TXT_FOLDER, SAVE_FOLDER, BG_EXPAND_RATIO)
        yolo_analyzer.process_all()

        # 步骤2：执行代码1，生成原始Excel（txt保存在IMAGE_FOLDER中，直接作为输入）
        print("\n" + "="*80)
        print("开始执行步骤2：解析缺陷txt → 生成原始Excel")
        print("="*80)
        excel_generator = DefectResultExcelGenerator(SAVE_FOLDER)
        original_excel_path = excel_generator.generate_excel(ORIGINAL_EXCEL_NAME)

        # 步骤3：执行代码2，生成带分级Excel
        print("\n" + "="*80)
        print("开始执行步骤3：原始Excel → 生成带分级Excel")
        print("="*80)
        graded_excel_path = os.path.join(SAVE_FOLDER, GRADED_EXCEL_NAME)
        process_excel_to_graded(original_excel_path, graded_excel_path, FACTORS_CONFIG, WEIGHTS)

        # 步骤4：执行代码3，生成等级标注txt
        print("\n" + "="*80)
        print("开始执行步骤4：带分级Excel → 生成等级标注txt")
        print("="*80)
        write_defect_severity_to_json(graded_excel_path, TXT_FOLDER, SAVE_FOLDER)

        # 流程完成
        print("\n" + "="*80)
        print("🎉 端到端流程全部执行完成！")
        print(f"1. 缺陷报告txt：保存在 {IMAGE_FOLDER} 中（每张图片对应一个txt）")
        print(f"2. 原始Excel：{original_excel_path}")
        print(f"3. 带分级Excel：{graded_excel_path}")
        print(f"4. 等级标注txt：{SAVE_FOLDER}")
        print("="*80)

    except Exception as e:
        print(f"\n❌ 流程执行失败：{str(e)}")